
import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
import calendar
from calendar import monthrange

st.set_page_config(page_title="CRMC Monthly Report Automation", layout="wide")
st.title("📊 CRMC Productivity Report Assistant")

# STEP 1: DUPLICATE SHEETS
st.header("Step 1: Create New Month Sheets")
uploaded_template = st.file_uploader("📎 Upload existing CRMC FY25 Production Report (.xlsx)", type="xlsx")
month_code = st.text_input("📅 Enter the new month code (e.g., Jun25)")

month_name = st.selectbox("📆 Select Month", list(calendar.month_name)[1:])
year_selected = st.number_input("🗓️ Enter Year (e.g., 2025)", min_value=2023, max_value=2100, value=2025)

def get_fiscal_info(month_name, year_selected):
    month_num = list(calendar.month_name).index(month_name)
    last_day = monthrange(year_selected, month_num)[1]
    fy_start_year = year_selected if month_name in ["November", "December"] else year_selected - 1

    b3 = f"For services November 1, {fy_start_year} through {month_name} {last_day}, {year_selected}"
    d6 = f"{year_selected}, {month_name}"
    e6 = f"{fy_start_year}, {month_name}"
    i6 = f"YTD {month_name} {year_selected}"
    j6 = f"YTP {month_name} {fy_start_year}"

    fiscal_months = ["November", "December", "January", "February", "March", "April", "May", "June", "July", "August", "September", "October"]
    divisor = fiscal_months.index(month_name) + 1 if month_name in fiscal_months else 1

    return b3, d6, e6, i6, j6, divisor

if uploaded_template and month_code:
    b3, d6, e6, i6, j6, divisor = get_fiscal_info(month_name, year_selected)
    wb_io = BytesIO(uploaded_template.read())
    wb = load_workbook(wb_io)
    sheet_names = wb.sheetnames
    selected_sheets = st.multiselect("✅ Select sheets to duplicate (e.g., April25_Primary)", options=sheet_names)

    if st.button("🚀 Create New Month Sheets"):
        created_sheets = []
        for original_name in selected_sheets:
            section = original_name.split("_")[-1] if "_" in original_name else original_name
            new_name = f"{month_code}_{section}"
            if new_name in wb.sheetnames:
                st.warning(f"Sheet '{new_name}' already exists. Skipping.")
                continue
            source_ws: Worksheet = wb[original_name]
            target_ws = wb.copy_worksheet(source_ws)
            target_ws.title = new_name
            for row in range(7, target_ws.max_row + 1):
                for col in range(4, 13):  # Clear D–L
                    target_ws.cell(row=row, column=col).value = None
            target_ws["B3"] = b3
            target_ws["D6"] = d6
            target_ws["E6"] = e6
            target_ws["I6"] = i6
            target_ws["J6"] = j6
            created_sheets.append(new_name)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        st.success(f"✅ Created and cleaned: {', '.join(created_sheets)}")
        st.download_button("⬇️ Download New Workbook", output, file_name="CRMC_NewMonth_Blank.xlsx")
        st.session_state["divisor"] = divisor

# STEP 2: POPULATE DATA
st.header("Step 2: Populate New Month Sheets")
app_file = st.file_uploader("📂 Upload APP Summary", type="xlsx")
clp_file = st.file_uploader("📂 Upload CLP Summary", type="xlsx")
misc_file = st.file_uploader("📂 Upload MISC Summary", type="xlsx")
psa_file = st.file_uploader("📂 Upload PSA Summary", type="xlsx")
template_file = st.file_uploader("📂 Upload Workbook with MonthYY Sheets (from Step 1)", type="xlsx", key="upload2")
month_code_step2 = st.text_input("🔁 Enter Month Code again for Step 2 (e.g., Jun25)")
manual_divisor = st.number_input("🔢 Divisor for Annualization", min_value=1, value=st.session_state.get("divisor", 6))

def clean_name(name):
    if not isinstance(name, str):
        return ""
    return (
        name.strip()
        .lower()
        .replace("\xa0", " ")
        .replace("\u200b", "")
        .replace("\n", " ")
        .replace("\r", "")
    )

def load_summary(file):
    df = pd.read_excel(file, sheet_name=0, skiprows=4, usecols="A:J")
    df = df.dropna(subset=[df.columns[0]])
    df.columns = [f"Col{i+1}" for i in range(len(df.columns))]
    df.rename(columns={"Col1": "Provider"}, inplace=True)
    df["CleanProvider"] = df["Provider"].map(clean_name)
    return df.set_index("CleanProvider")

if st.button("📤 Generate Final Report"):
    if not all([app_file, clp_file, misc_file, psa_file, template_file, month_code_step2]):
        st.error("Please upload all files and enter the month code.")
    else:
        with st.spinner("Processing and populating..."):
            df_app = load_summary(app_file)
            df_clp = load_summary(clp_file)
            df_misc = load_summary(misc_file)
            df_psa = load_summary(psa_file)

            df_primary = pd.concat([df_app, df_clp]).drop_duplicates()
            df_psa_combined = pd.concat([df_psa, df_app]).drop_duplicates()
            df_misc_combined = pd.concat([df_misc, df_app]).drop_duplicates()

            source_data = {
                "Primary": df_primary,
                "PSA": df_psa_combined,
                "MISC": df_misc_combined
            }

            unmatched_providers = []
            wb = load_workbook(BytesIO(template_file.read()))

            for section, df in source_data.items():
                sheet_name = f"{month_code_step2}_{section}"
                if sheet_name not in wb.sheetnames:
                    st.warning(f"⚠️ Sheet '{sheet_name}' not found. Skipping.")
                    continue

                ws = wb[sheet_name]
                row = 7
                while True:
                    provider = ws[f"B{row}"].value
                    if provider is None:
                        break
                    cleaned_provider = clean_name(provider)
                    if cleaned_provider in df.index:
                        for col_offset in range(1, 10):
                            col_letter = get_column_letter(3 + col_offset)
                            try:
                                value = df.iloc[df.index.get_loc(cleaned_provider), col_offset]
                                ws[f"{col_letter}{row}"] = value
                            except Exception as e:
                                unmatched_providers.append((provider, section, str(e)))
                        ws.cell(row=row, column=15).value = f"=I{row}/{manual_divisor}*12"
                    else:
                        unmatched_providers.append((provider, section, "No match"))
                    row += 1

            output = BytesIO()
            wb.save(output)
            output.seek(0)
            st.success("✅ Final report generated!")
            st.download_button("📥 Download Completed Report", output, file_name=f"CRMC_Populated_{month_code_step2}.xlsx")

            if unmatched_providers:
                st.warning("⚠️ Some providers were not matched.")
                unmatched_df = pd.DataFrame(unmatched_providers, columns=["Provider", "Section", "Issue"])
                st.dataframe(unmatched_df)
                unmatched_xlsx = BytesIO()
                unmatched_df.to_excel(unmatched_xlsx, index=False)
                unmatched_xlsx.seek(0)
                st.download_button("⬇️ Download Unmatched Providers", unmatched_xlsx, file_name="Unmatched_Providers.xlsx")
